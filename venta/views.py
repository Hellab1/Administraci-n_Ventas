from django.http.response import Http404
from django.shortcuts import render
# Instanciamos las vistas genéricas de Django 
from django.views.generic import ListView, DetailView 
from django.views.generic.edit import CreateView, UpdateView, DeleteView
# Instanciamos los modelos para poder usarlo en nuestras Vistas CRUD
from .models import Cliente, Detalle_Venta, Factura, Producto, Factura
# Nos sirve para redireccionar despues de una acción revertiendo patrones de expresiones regulares 
from django.urls import reverse 
# Habilitamos el uso de mensajes en Django
from django.contrib import messages  
# Habilitamos los mensajes para class-based views 
from django.contrib.messages.views import SuccessMessageMixin  
# Habilitamos los formularios en Django
from django import forms
from django.contrib.auth.mixins import LoginRequiredMixin

from rest_framework import generics
from .serializers import detallesSerializer, ventaSerializer

from django.shortcuts import redirect
from .forms import *
from django.contrib.auth.decorators import login_required

# Create your views here.
class ClienteListado(LoginRequiredMixin, ListView): 
    model = Cliente # Llamamos a la clase 'Cliente' que se encuentra en nuestro archivo 'models.py'
    login_url = '/iniciar-sesion/'
class ClienteCrear(LoginRequiredMixin, SuccessMessageMixin, CreateView): 
    model = Cliente # Llamamos a la clase 'Cliente' que se encuentra en nuestro archivo 'models.py'
    form = Cliente # Definimos nuestro formulario con el nombre de la clase o modelo 'Cliente'
    fields = "__all__" # Le decimos a Django que muestre todos los campos de la tabla 'Cliente' de nuestra Base de Datos 
    success_message = 'Cliente Creado Correctamente !' # Mostramos este Mensaje luego de Crear un Cliente
    login_url = '/iniciar-sesion/'
    # Redireccionamos a la página principal luego de crear un registro o Cliente
    def get_success_url(self):        
        return reverse('leer_cliente') # Redireccionamos a la vista principal 'leer'

"""class ClienteDetalle(LoginRequiredMixin, DetailView): 
    model = Cliente # Llamamos a la clase 'Cliente' que se encuentra en nuestro archivo 'models.py'
    login_url = '/iniciar-sesion/' """

class ClienteActualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView): 
    model = Cliente # Llamamos a la clase 'Cliente' que se encuentra en nuestro archivo 'models.py' 
    form = Cliente # Definimos nuestro formulario con el nombre de la clase o modelo 'Cliente' 
    fields = "__all__" # Le decimos a Django que muestre todos los campos de la tabla 'Cliente' de nuestra Base de Datos 
    success_message = 'Cliente Actualizado Correctamente !' # Mostramos este Mensaje luego de Editar un Cliente 
    login_url = '/iniciar-sesion/' 
    # Redireccionamos a la página principal luego de actualizar un registro o Cliente
    def get_success_url(self):               
        return reverse('leer_cliente') # Redireccionamos a la vista principal 'leer'

class ClienteEliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView): 
    model = Cliente 
    form = Cliente
    fields = "__all__"     
    login_url = '/iniciar-sesion/' 
    # Redireccionamos a la página principal luego de eliminar un registro o Cliente
    def get_success_url(self): 
        success_message = 'Cliente Eliminado Correctamente !' # Mostramos este Mensaje luego de Editar un Cliente 
        messages.success (self.request, (success_message))       
        return reverse('leer_cliente') # Redireccionamos a la vista principal 'leer'

class ProductoListado(LoginRequiredMixin, ListView): 
    model = Producto
    login_url = '/iniciar-sesion/'

class ProductoCrear(LoginRequiredMixin, SuccessMessageMixin, CreateView): 
    model = Producto
    form = Producto
    fields = "__all__" 
    success_message = 'Producto Creado Correctamente !' 
    login_url = '/iniciar-sesion/'
    def get_success_url(self):        
        return reverse('leer_producto') 

class ProductoActualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView): 
    model = Producto 
    form = Producto
    fields = "__all__"
    success_message = 'Producto Actualizado Correctamente !' 
    login_url = '/iniciar-sesion/'
    def get_success_url(self):               
        return reverse('leer_producto') 

class ProductoEliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView): 
    model = Producto 
    form = Producto
    fields = "__all__"     
    login_url = '/iniciar-sesion/'
    def get_success_url(self): 
        success_message = 'Producto Eliminado Correctamente !'
        messages.success (self.request, (success_message))       
        return reverse('leer_producto') 

class OTListado(LoginRequiredMixin, ListView): 
    model = Factura
    login_url = '/iniciar-sesion/'

"""
class OTCrear(LoginRequiredMixin, SuccessMessageMixin, CreateView): 
    model = Orden_Compra
    form = Orden_Compra
    fields = "__all__" 
    success_message = 'Orden de Compra Creada Correctamente!' 
    login_url = '/iniciar-sesion/'
    def get_success_url(self):        
        return reverse('leer_ot') 

class OTDetalle(LoginRequiredMixin, DetailView): 
    model = Orden_Compra
    login_url = '/iniciar-sesion/'
"""

class OTActualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView): 
    model = Factura 
    form = Factura
    fields = "__all__"
    success_message = 'Orden de Compra Actualizada Correctamente !' 
    login_url = '/iniciar-sesion/'
    def get_success_url(self):               
        return reverse('leer_ot') 

class OTEliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView): 
    model = Factura 
    form = Factura
    fields = "__all__"   
    login_url = '/iniciar-sesion/' 
    def get_success_url(self): 
        success_message = 'Orden de Compra Eliminada Correctamente !'
        messages.success (self.request, (success_message))       
        return reverse('leer_ot') 

class ventasList(generics.ListAPIView):
    queryset = Factura.objects.all()
    serializer_class = ventaSerializer

class detallesList(generics.ListAPIView):
    queryset = Factura.objects.all()
    serializer_class = detallesSerializer

@login_required
def facturaCrear(request):    
    if request.method == "POST":
        form = facturaForm(request.POST)
        if form.is_valid():            
            orden = form.save(commit=False) 
            orden.neto = 0
            orden.iva = 0
            orden.total = 0
            orden.save()
            return redirect('detalles', pk=orden.id_factura)
    else:
        form = facturaForm()
    return render(request, 'venta/crear_ot.html', {'form': form})   

@login_required
def detallesCrearVenta(request, pk):
    factura = Factura.objects.get(id_factura=pk)
    details = Detalle_Venta.objects.filter(orden_venta=pk)
    id = pk
    if request.method == "POST" and 'btnCrear' in request.POST:
        id_producto = request.POST['producto']
        producto = Producto.objects.get(id_producto=id_producto)
        form = DetalleForm(request.POST)
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.total_detalle = producto.precio_neto*detalle.cantidad
            detalle.paquetes = 0
            if producto.espesor == 9:
                detalle.piezas_paquete = 100
            elif producto.espesor == 12:
                detalle.piezas_paquete = 90
            elif producto.espesor == 15:
                detalle.piezas_paquete = 80
            else:
                detalle.piezas_paquete = 0
            detalle.orden_venta_id = id
            detalle.save()
            factura.neto = factura.neto + detalle.total_detalle
            factura.iva = factura.iva + round(detalle.total_detalle * 0.19)
            factura.total = factura.total + round(detalle.total_detalle * 1.19)
            factura.save()
            return redirect('detalles', id)
    else:
        form = DetalleForm()
    return render(request, 'venta/detalles_venta.html', {'form': form, 'details':details})

class indexFactura(LoginRequiredMixin, ListView): 
    model = Factura
    login_url = '/iniciar-sesion/'

# from django.template import RequestContext
from django.shortcuts import render_to_response
 
#404: página no encontrada
def pag_404_not_found(request, exception, template_name="venta/404.html"):
    response = render_to_response("venta/404.html")
    response.status_code=404
    return response

#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
 
#Nuestra clase hereda de la vista genérica TemplateView
class ReporteFacturasExcel(TemplateView):     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        facturas = Factura.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE FACTURAS'
        #ws['B1'] = 'REPORTE DE FACTURAS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        #ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A1'] = 'ID'
        ws['B1'] = 'CODIGO'
        ws['C1'] = 'NUMERO DE PEDIDO'
        ws['D1'] = 'FECHA' 
        ws['E1'] = 'RUT' 
        ws['F1'] = 'NOMBRE CLIENTE'
        ws['G1'] = 'DIRECCION'
        ws['H1'] = 'TIPO DE PRODUCTO'
        ws['I1'] = 'DESCRIPCION' 
        ws['J1'] = 'CANTIDAD'
        ws['K1'] = 'PAQUETE'
        ws['L1'] = 'PIEZAS POR PAQUETE'
        ws['M1'] = 'NETO'
        ws['N1'] = 'IVA' 
        ws['O1'] = 'TOTAL'
        ws['P1'] = 'TIPO DE PAGO'
        ws['Q1'] = 'FORMA DE PAGO'
        ws['R1'] = 'TIPO DE FACTURACIÓN' 
    
        cont=2
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for factura in facturas:
            ws.cell(row=cont,column=1).value = factura.id_factura
            ws.cell(row=cont,column=2).value = factura.codigo
            ws.cell(row=cont,column=3).value = factura.n_pedido
            ws.cell(row=cont,column=4).value = factura.fecha
            ws.cell(row=cont,column=5).value = factura.rut
            ws.cell(row=cont,column=6).value = factura.nombre
            ws.cell(row=cont,column=7).value = factura.direccion
            ws.cell(row=cont,column=8).value = factura.tipo
            ws.cell(row=cont,column=9).value = factura.descripcion_productos
            ws.cell(row=cont,column=10).value = factura.cantidad
            ws.cell(row=cont,column=11).value = factura.paquete
            ws.cell(row=cont,column=12).value = factura.piezas_paquete
            ws.cell(row=cont,column=13).value = factura.neto
            ws.cell(row=cont,column=14).value = factura.iva
            ws.cell(row=cont,column=15).value = factura.total
            ws.cell(row=cont,column=16).value = factura.tipo_pago
            ws.cell(row=cont,column=17).value = factura.forma_pago
            ws.cell(row=cont,column=18).value = factura.tipo_facturacion
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReportefacturasExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response